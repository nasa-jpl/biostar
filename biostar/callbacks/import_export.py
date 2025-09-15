import io
import time
from copy import copy
from datetime import datetime
from pathlib import Path

import dash_mantine_components as dmc
from dash import Dash, dcc, html
from dash.dependencies import Input, Output, State
from dash.exceptions import PreventUpdate
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from biostar.modules.data import (
    POSTERIOR_MAP,
    PPEL_TABLE_COLUMNS,
    PPEL_TABLE_METADATA_COLUMNS,
    SAMPLE_TABLE_COLUMNS,
    SAMPLE_TABLE_METADATA_COLUMNS,
    SPEC_ISO_CLASSES,
    find_by_key,
    parse_ppel_upload,
    parse_pps_upload,
)


def validate_hardware_joint(hardware_list: list[dict]) -> list:
    """Check to make sure the provided hardware elements are (jointly) valid"""

    def implied_level(hw):
        if hw["parent_id"] is None:
            return 2
        parent = find_by_key(hardware_list, "id", hw["parent_id"])
        if parent is None:
            return 0
        return implied_level(parent) + 1

    # Ensure all parent IDs reference a valid hardware ID
    hw_ids = [hw["id"] for hw in hardware_list]
    if not all([not hw["parent_id"] or hw["parent_id"] in hw_ids for hw in hardware_list]):
        return False, hardware_list

    # Ensure all hardware serving as parent have is_component False
    parent_ids = [hw["parent_id"] for hw in hardware_list]
    if not all(
        [
            not hw["is_component"] if hw["id"] in parent_ids else hw["is_component"]
            for hw in hardware_list
        ]
    ):
        return False, hardware_list

    # Ensure all level labels are correct
    if not all([implied_level(hw) == hw["level"] for hw in hardware_list]):
        return False, hardware_list

    # Determine/assign validity of components
    for hw in sorted(
        hardware_list,
        key=lambda x: {None: 0, "Sampled": 1, "Unsampled - Spec": 2, "Unsampled - Implied": 3}[
            x["type"]
        ],
    ):
        comp_valid = False
        if hw["is_component"] and hw["dim"]:
            comp_valid = all(
                [hw["area"] if hw["dim"].startswith("2") else hw["volume"], hw["type"]]
            )
            if comp_valid and hw["type"] == "Sampled":
                comp_valid = hw["analogy"] in POSTERIOR_MAP
            elif hw["type"] == "Unsampled - Implied":
                implied_hw = find_by_key(hardware_list, "id", hw["implied_id"])
                comp_valid = (implied_hw is not None) and implied_hw["valid"]
            elif hw["type"] == "Unsampled - Spec":
                comp_valid = hw["spec"] in SPEC_ISO_CLASSES[hw["dim"]]
        tgt_hw = find_by_key(hardware_list, "id", hw["id"])
        tgt_hw["valid"] = comp_valid

    return True, hardware_list


def validate_sample(sample: dict, hardware_dict: dict, samples_list: list[dict]):
    """Check to make sure the provided sample is valid within the current context of the project"""

    hw_id = sample["Hardware ID"]
    samp_id = sample["Sample ID"]

    if (not all([hw_id, samp_id])) or (hw_id not in hardware_dict):
        return samp_id, "inv_id"

    if not hardware_dict[hw_id]["is_component"] or not all(
        [hardware_dict[hw_id]["type"] == "Sampled", hardware_dict[hw_id]["valid"]]
    ):
        return samp_id, "inv_component"

    if len([s for s in samples_list if s["Sample ID"] == samp_id]) > 1:
        return samp_id, "inv_duplicate"

    return sample


def attach_callbacks(app: Dash):
    """"""

    @app.callback(
        Output("download-ppel-export", "data"),
        Input("button-ppel-export", "n_clicks"),
        State("datatable-ppel", "data"),
        State("datatable-samples", "data"),
        prevent_initial_call=True,
    )
    def export_ppel(_, ppel_rows: list[dict], sample_rows: list[dict]):
        """Download PPEL to Excel file, including sample info"""

        # Create output file and load template file
        pseudofile = io.BytesIO()
        wb = load_workbook(
            Path(__file__).parent.parent / "static/biostar/templates/template_ppel.xlsx"
        )
        ws_colors = wb["Color_codes"]

        def get_ppel_row_mode(data):
            if data["Level"] == 1:
                return "project"
            elif data["Level"] == 2 and data["Hardware Type"] == "Rollup":
                return "rollup_noparent"
            elif data["Level"] == 2:
                return "component_noparent"
            elif data["Hardware Type"] == "Rollup":
                return "rollup"
            return "component"

        def get_style_source_ppel(col, row_mode="component"):
            if (
                row_mode == "project"
                and col
                in ["Parent ID", "Dimensionality", "Analogy", "Origin", "Spec Class"]
                + PPEL_TABLE_METADATA_COLUMNS
            ):
                return ws_colors.cell(row=4, column=1)
            elif row_mode == "rollup_noparent" and col in [
                "Parent ID",
                "Dimensionality",
                "Analogy",
                "Origin",
                "Spec Class",
            ]:
                return ws_colors.cell(row=4, column=1)
            elif row_mode == "component_noparent" and col in ["Parent ID"]:
                return ws_colors.cell(row=4, column=1)
            elif row_mode == "rollup" and col in [
                "Dimensionality",
                "Analogy",
                "Origin",
                "Spec Class",
            ]:
                return ws_colors.cell(row=4, column=1)
            elif col in PPEL_TABLE_METADATA_COLUMNS:
                return ws_colors.cell(row=2, column=1)
            elif col in [
                "Sampled Area",
                "Sampled Volume",
                "Spec Value",
                "CBE Bioburden Density (2D)",
                "CBE Bioburden Density (3D)",
                "CBE Spore Bioburden (2D)",
                "CBE Spore Bioburden (3D)",
            ]:
                return ws_colors.cell(row=3, column=1)
            return ws_colors.cell(row=1, column=1)

        def get_style_source_samples(col, _):
            if col in SAMPLE_TABLE_METADATA_COLUMNS:
                return ws_colors.cell(row=2, column=1)
            return ws_colors.cell(row=1, column=1)

        # Iterate over sheets we need to edit
        ppel_cols = [x["name"] for x in PPEL_TABLE_COLUMNS]
        sample_cols = [x["name"] for x in SAMPLE_TABLE_COLUMNS]
        for rows, cols, sheet in zip(
            [ppel_rows, sample_rows], [ppel_cols, sample_cols], ["PPEL", "Samples"]
        ):
            ws = wb[sheet]
            style_source_fn = get_style_source_ppel if sheet == "PPEL" else get_style_source_samples

            # Write new data
            for row_num, data in enumerate(rows, start=2):
                for k, v in data.items():
                    cell = ws.cell(row=row_num, column=(cols.index(k) + 1))
                    style_source = style_source_fn(
                        k, get_ppel_row_mode(data) if sheet == "PPEL" else None
                    )
                    cell.fill = copy(style_source.fill)
                    cell.value = v

            # Clear any remaining dummy data
            for row in ws.iter_rows(min_row=(2 + len(rows)), max_row=ws.max_row):
                for cell in row:
                    cell.value = None
                    cell.fill = PatternFill()

        # Save to pseudofile
        wb.save(pseudofile)
        pseudofile.seek(0)

        return dcc.send_bytes(
            pseudofile.getvalue(), f"ppel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    @app.callback(
        Output("modal-import-hardware", "opened", allow_duplicate=True),
        Input("button-import-hardware", "n_clicks"),
        prevent_initial_call=True,
    )
    def import_ppel_open(_):
        """Open the modal to import samples from PPS format"""

        return True

    @app.callback(
        Output("modal-import-hardware", "opened", allow_duplicate=True),
        Input("button-import-hardware-close", "n_clicks"),
        prevent_initial_call=True,
    )
    def import_ppel_close(_):
        """Open the modal to import samples from PPS format"""

        return False

    @app.callback(
        Output("datatable-groups", "data", allow_duplicate=True),
        Output("hardware-json", "data", allow_duplicate=True),
        Output("datatable-samples", "data", allow_duplicate=True),
        Output("project-json", "data", allow_duplicate=True),
        Output("ppel-storage", "data", allow_duplicate=True),
        Output("ppel-wipe-flag", "data"),
        Output("select-results-hardware-id", "value"),
        Output("filename-import-hardware", "children"),
        Output("warnings-import-hardware", "children"),
        Input("upload-import-hardware", "contents"),
        State("upload-import-hardware", "filename"),
        State("datatable-groups", "data"),
        State("hardware-json", "data"),
        State("datatable-samples", "data"),
        State("project-json", "data"),
        State("ppel-wipe-flag", "data"),
        State("select-results-hardware-id", "value"),
        prevent_initial_call=True,
    )
    def import_ppel_stage_1(
        contents: str,
        filename: str,
        groups_list: list[dict],
        hardware_dict: dict,
        samples_list: list[dict],
        project_dict: dict,
        wipe_idx: int,
        results_hw_id: str,
    ):
        """Parse uploaded PPEL file, validate hardware & samples, wipe current state and queue new state for import"""

        hw_type_map = {
            "Rollup": None,
            "Unsampled - Implied": "Unsampled - Implied",
            "Unsampled - Spec": "Unsampled - Spec",
            "Sampled": "Sampled",
        }

        def ppel_to_biostar(row: dict):
            return {
                "id": row["Hardware ID"],
                "parent_id": row["Parent ID"],
                "level": row["Level"],
                "group": row["Grouping"],
                "is_component": row["Hardware Type"] != "Rollup",
                "valid": False,  # components will be checked for validity below
                "dim": row["Dimensionality"] if row["Dimensionality"] else "2D Area",
                "area": row["Total Area"] if row["Total Area"] else "",
                "volume": row["Total Volume"] if row["Total Volume"] else "",
                "type": hw_type_map[row["Hardware Type"]],
                "analogy": row["Analogy"],
                "implied_id": row["Origin"],
                "spec": row["Spec Class"],
                "handling": row["Handling Constraints"],
                "ventilation": row["Ventilation"],
                "composition": row["Material Composition"],
                "cleaning_fab": row["Cleaning Procedures (Fabrication)"],
                "cleaning_pre": row["Cleaning Procedures (Pre-SI&T)"],
                "cleaning_sit": row["Cleaning Procedures (SI&T)"],
                "reduction_fab": row["Bioburden Reduction (Fabrication)"],
                "reduction_pre": row["Bioburden Reduction (Pre-SI&T)"],
                "reduction_sit": row["Bioburden Reduction (SI&T)"],
                "notes": row["Hardware Notes"],
            }

        # Parse the uploaded filename and contents
        last_upload_str = f"Last Upload: '{filename}'"
        success, result = parse_ppel_upload(contents, filename)

        # Report error if we could not parse the upload
        if not success:
            warning = dmc.Alert(
                result,
                title="Error!",
                color="red",
                mt=16,
            )
            return (
                groups_list,
                hardware_dict,
                samples_list,
                project_dict,
                {},
                wipe_idx,
                results_hw_id,
                last_upload_str,
                warning,
            )
        (ppel_rows, ppel_unresolved, sample_rows, sample_unresolved, project_dict_new) = result

        # Extract the hardware elements and filter for valid IDs
        hardware_list_new = [ppel_to_biostar(row) for row in ppel_rows]
        hardware_list_new = [hw for hw in hardware_list_new if hw["id"]]
        alerts_hw = {"inv_id": len(ppel_rows) - len(hardware_list_new)}

        # Validate the hierarchy in the set of hardware
        # Also assigns `valid` flag to each individual hardware element
        # Report error if we could not parse the hierarchy
        valid, hardware_list_new = validate_hardware_joint(hardware_list_new)
        if not valid:
            if not ppel_unresolved:
                warning = dmc.Alert(
                    "Error resolving project hierarchy - no categorical variable warnings, so ensure all parent-child relationships are logical (including hardware levels).",
                    title="Error!",
                    color="red",
                    my=16,
                )
            else:
                warning = dmc.Alert(
                    f"Error resolving project hierarchy - failed to resolve categorical variables for the following cases: {ppel_unresolved}.",
                    title="Error!",
                    color="red",
                    my=16,
                )
            return (
                groups_list,
                hardware_dict,
                samples_list,
                project_dict,
                {},
                wipe_idx,
                results_hw_id,
                last_upload_str,
                warning,
            )
        hardware_dict_new = {hw["id"]: hw for hw in hardware_list_new}

        # Extract samples and filter for valid IDs
        samples_list_new = [validate_sample(s, hardware_dict_new, sample_rows) for s in sample_rows]
        alerts_samp_raw = [s for s in samples_list_new if isinstance(s, tuple)]
        alerts_samp = {
            "inv_id": [
                samp_id for (samp_id, alert_code) in alerts_samp_raw if alert_code == "inv_id"
            ],
            "inv_component": [
                samp_id
                for (samp_id, alert_code) in alerts_samp_raw
                if alert_code == "inv_component"
            ],
            "inv_duplicate": [
                samp_id
                for (samp_id, alert_code) in alerts_samp_raw
                if alert_code == "inv_duplicate"
            ],
        }
        samples_list_new = [s for s in samples_list_new if isinstance(s, dict)]

        # Construct updated groups
        group_tags_new = set([hw["group"] for hw in hardware_list_new if hw["group"]])
        groups_new = [
            {
                "Group Tag": tag,
                "Target Density (2D)": find_by_key(ppel_rows, "Grouping", tag)[
                    "Grouping Target Density (2D)"
                ],
                "Target Density (3D)": find_by_key(ppel_rows, "Grouping", tag)[
                    "Grouping Target Density (3D)"
                ],
            }
            for tag in group_tags_new
        ]

        # Generate info pop-up
        warnings = []
        if ppel_unresolved:
            warnings.append(
                f"Failed to resolve categorical variables for some hardware, replacing with empty string. Affected cases: {ppel_unresolved}"
            )
        if sample_unresolved:
            warnings.append(
                f"Failed to resolve categorical variables for some samples, replacing with empty string. Affected cases: {sample_unresolved}"
            )
        if alerts_hw["inv_id"]:
            warnings.append(
                f"Skipped importing {alerts_hw['inv_id']} elements due to missing Hardware ID"
            )
        if alerts_samp["inv_id"]:
            warnings.append(
                f"Skipped importing the following sample(s) due to missing/invalid Sample ID or Hardware ID: {alerts_samp['inv_id']}"
            )
        if alerts_samp["inv_component"]:
            warnings.append(
                f"Skipped importing the following sample(s) with Hardware ID of an invalid element (not a component, or not properly configured with Component Type = 'Sampled'): {alerts_samp['inv_component']}"
            )
        if alerts_samp["inv_duplicate"]:
            warnings.append(
                f"Skipped importing {len(alerts_samp['inv_duplicate'])} sample(s) with duplicate Sample ID"
            )
        infos = []
        if hardware_dict:
            infos.append(
                f"Overwrote {len(hardware_dict)} old hardware element(s) that existed prior to upload"
            )
        if samples_list:
            infos.append(
                f"Overwrote {len(samples_list)} old sample(s) that existed prior to upload"
            )
        if groups_list:
            infos.append(f"Overwrote {len(groups_list)} old groups(s) that existed prior to upload")
        alert = dmc.Alert(
            [
                f"Successfully imported PPEL [ +{len(hardware_dict_new)} hardware element(s), +{len(samples_list_new)} sample(s), +{len(groups_new)} group(s) ].",
                html.Ul(
                    [html.Li(m) for m in infos + warnings],
                    style={"marginBottom": 0, "marginTop": 8},
                ),
            ],
            mt=16,
            **(
                {"title": "Success", "color": "green"}
                if not warnings
                else {"title": "Warning", "color": "yellow"}
            ),
        )

        return (
            [],
            {},
            [],
            {"name": "My Project", "group": ""},
            {
                "groups": groups_new,
                "hardware": hardware_dict_new,
                "samples": samples_list_new,
                "project": project_dict_new,
            },
            wipe_idx + 1,
            "",
            last_upload_str,
            alert,
        )

    @app.callback(
        Output("datatable-groups", "data", allow_duplicate=True),
        Output("hardware-json", "data", allow_duplicate=True),
        Output("datatable-samples", "data", allow_duplicate=True),
        Output("project-json", "data", allow_duplicate=True),
        Output("ppel-storage", "data", allow_duplicate=True),
        Output("ppel-import-flag", "data"),
        Input("ppel-wipe-flag", "data"),
        State("ppel-import-flag", "data"),
        State("ppel-storage", "data"),
        prevent_initial_call=True,
    )
    def import_ppel_stage_2(wipe_idx: int, import_idx: int, ppel_storage: dict):
        """Import queued tool state following wipe stage of PPEL import"""

        # Only need to trigger if the wipe_idx has increased past import_idx
        if wipe_idx == import_idx:
            raise PreventUpdate

        # Sleep to give wipe time stage time to process
        time.sleep(1)

        return (
            ppel_storage["groups"],
            ppel_storage["hardware"],
            ppel_storage["samples"],
            ppel_storage["project"],
            {},
            import_idx + 1,
        )

    @app.callback(
        Output("modal-import-samples", "opened", allow_duplicate=True),
        Input("button-import-samples", "n_clicks"),
        prevent_initial_call=True,
    )
    def import_pps_open(_):
        """Open the modal to import samples from PPS format"""

        return True

    @app.callback(
        Output("modal-import-samples", "opened", allow_duplicate=True),
        Input("button-import-samples-close", "n_clicks"),
        prevent_initial_call=True,
    )
    def import_pps_close(_):
        """Open the modal to import samples from PPS format"""

        return False

    @app.callback(
        Output("datatable-samples", "data", allow_duplicate=True),
        Output("filename-import-samples", "children"),
        Output("warnings-import-samples", "children"),
        Input("upload-import-samples", "contents"),
        State("upload-import-samples", "filename"),
        State("hardware-json", "data"),
        State("datatable-samples", "data"),
        prevent_initial_call=True,
    )
    def import_pps(contents: str, filename: str, hardware_dict: dict, samples_list: list[dict]):
        """Parse uploaded PPS file, validate rows, and update sampling storage"""

        def pps_to_biostar(row: dict):
            hw_id = row["Zone"]
            return {
                "Sample ID": row["Sample Number"],
                "Hardware ID": hw_id,
                "PP Accountable": row["PP Accountable"],
                "Sampled Area": None,
                "Sampled Volume": None,
                "Sampling Device": row["Sampling Method"],
                "Sampling Device Type": "Puritan Cotton"
                if row["Sampling Method"] == "Swab"
                else "TX3211",
                "Processing Technique": "NASA Standard",
                "Pour Fraction": 1,
                "CFU": int(row["72 Count"]),
                "Assay Name": row["Assay Name"],
                "Assay Date": row["Assay Date"],
                "PP Cert #": row["PP Cert #"],
                "Control Type": row["Control Type"],
                "Sampling Notes": row["Sampling Notes"],
            }

        # Parse the uploaded filename and contents
        last_upload_str = f"Last Upload: '{filename}'"
        success, result = parse_pps_upload(contents, filename)

        # Report errors if we could not parse the upload
        if not success:
            warning = dmc.Alert(
                result,
                title="Error!",
                color="red",
                mt=16,
            )
            return samples_list, last_upload_str, warning
        (sample_rows, sample_unresolved) = result

        # Extract and count the valid/invalid rows
        samples_list_new = [pps_to_biostar(row) for row in sample_rows]
        samples_list_new = [
            validate_sample(s, hardware_dict, samples_list + samples_list_new)
            for s in samples_list_new
        ]
        alerts_samp_raw = [s for s in samples_list_new if isinstance(s, tuple)]
        alerts_samp = {
            "inv_id": [
                samp_id for (samp_id, alert_code) in alerts_samp_raw if alert_code == "inv_id"
            ],
            "inv_component": [
                samp_id
                for (samp_id, alert_code) in alerts_samp_raw
                if alert_code == "inv_component"
            ],
            "inv_duplicate": [
                samp_id
                for (samp_id, alert_code) in alerts_samp_raw
                if alert_code == "inv_duplicate"
            ],
        }
        samples_list_new = [s for s in samples_list_new if isinstance(s, dict)]

        # Resolve area/volume now that we only have validated samples
        for samp in samples_list_new:
            row = find_by_key(sample_rows, "Sample Number", samp["Sample ID"])
            samp["Sampled Area"] = (
                float(row["Raw Area Sampled"])
                if hardware_dict[samp["Hardware ID"]]["dim"].startswith("2")
                else None
            )
            samp["Sampled Volume"] = (
                float(row["Raw Area Sampled"])
                if hardware_dict[samp["Hardware ID"]]["dim"].startswith("3")
                else None
            )

        # Generate alert if needed
        warnings = []
        if sample_unresolved:
            warnings.append(
                f"Failed to resolve categorical variables for some samples, replacing with empty string. Affected cases: {sample_unresolved}"
            )
        if alerts_samp["inv_id"]:
            warnings.append(
                f"Skipped importing the following sample(s) due to missing/invalid Sample ID or Hardware ID: {alerts_samp['inv_id']}"
            )
        if alerts_samp["inv_component"]:
            warnings.append(
                f"Skipped importing the following sample(s) with Hardware ID of an invalid element (not a component, or not properly configured with Component Type = 'Sampled'): {alerts_samp['inv_component']}"
            )
        if alerts_samp["inv_duplicate"]:
            warnings.append(
                f"Skipped importing {len(alerts_samp['inv_duplicate'])} sample(s) with duplicate Sample ID"
            )
        alert = dmc.Alert(
            [
                f"Successfully imported PPS [ +{len(samples_list_new)} sample(s) ].",
                html.Ul([html.Li(w) for w in warnings], style={"marginBottom": 0, "marginTop": 8}),
            ],
            mt=16,
            **(
                {"title": "Success", "color": "green"}
                if not warnings
                else {"title": "Warning", "color": "yellow"}
            ),
        )

        return samples_list + samples_list_new, last_upload_str, alert
