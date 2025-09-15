import base64
import io
import json
import os
from collections import defaultdict
from datetime import datetime
from itertools import product

import numpy as np
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook

_ = load_dotenv()

PKG_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PATH_HIERARCHY = os.getenv("PATH_HIERARCHY", os.path.join(PKG_DIR, "data/hierarchy_default.xlsx"))

PATH_POSTERIOR = os.getenv("PATH_POSTERIOR", os.path.join(PKG_DIR, "data/posterior_default.json"))

PATH_SEMANTIC_MAP = os.getenv(
    "PATH_SEMANTIC_MAP", os.path.join(PKG_DIR, "data/semantic_map_default.json")
)

PATH_METADATA = os.getenv("PATH_METADATA", os.path.join(PKG_DIR, "data/metadata_default.json"))

COMPONENT_TYPES = [
    "Sampled",
    "Unsampled - Implied",
    "Unsampled - Spec",
]

PROCESSING_TECHNIQUES = [
    "NASA Standard",
    "NASA Standard (w/ Membrane Filtration)",
    "ESA Standard",
    "ESA Standard (w/ Membrane Filtration)",
]

SPEC_ISO_CLASSES = {
    "2D (Area)": [
        "Surface, ISO<=7, Bio Control",
        "Surface, ISO<=7, Particle Control",
        "Surface, ISO=8, Bio Control",
        "Surface, ISO=8, Particle Control",
        "Surface, Uncontrolled",
        "Enclosed Surface, Cleanroom, Particle & Bio Control",
        "Enclosed Surface, Cleanroom, Only Particle Control",
        "Enclosed Surface, Uncontrolled Manufacturing",
    ],
    "3D (Volume)": [
        "Encapsulated, Electronics Piece Parts",
        "Encapsulated, Non-metal Avg",
        "Encapsulated, Non-metal Other",
    ],
}

SAMPLING_DEVICE_TYPE_MAP = {
    "Swab": ["Puritan Cotton", "Nylon-flocked", "Copan Polyester", "Copan Cotton"],
    "Wipe": ["TX3211", "TX3224"],
}

SAMPLE_TABLE_COLUMNS = [
    {"name": "Sample ID", "id": "Sample ID", "hideable": True, "type": "text", "editable": False},
    {
        "name": "Hardware ID",
        "id": "Hardware ID",
        "hideable": True,
        "type": "text",
        "editable": False,
    },
    {
        "name": "PP Accountable",
        "id": "PP Accountable",
        "hideable": True,
        "type": "text",
        "editable": True,
    },
    {
        "name": "Sampled Area",
        "id": "Sampled Area",
        "hideable": True,
        "type": "numeric",
        "editable": True,
    },
    {
        "name": "Sampled Volume",
        "id": "Sampled Volume",
        "hideable": True,
        "type": "numeric",
        "editable": True,
    },
    {
        "name": "Sampling Device",
        "id": "Sampling Device",
        "hideable": True,
        "type": "text",
        "editable": True,
    },
    {
        "name": "Sampling Device Type",
        "id": "Sampling Device Type",
        "hideable": True,
        "type": "text",
        "editable": True,
    },
    {
        "name": "Processing Technique",
        "id": "Processing Technique",
        "hideable": True,
        "type": "text",
        "editable": True,
    },
    {
        "name": "Pour Fraction",
        "id": "Pour Fraction",
        "hideable": True,
        "type": "numeric",
        "editable": True,
    },
    {"name": "CFU", "id": "CFU", "hideable": True, "type": "numeric", "editable": True},
    {
        "name": "Assay Name",
        "id": "Assay Name",
        "hideable": True,
        "type": "text",
        "editable": True,
    },
    {
        "name": "Assay Date",
        "id": "Assay Date",
        "hideable": True,
        "type": "datetime",
        "editable": True,
    },
    {
        "name": "PP Cert #",
        "id": "PP Cert #",
        "hideable": True,
        "type": "text",
        "editable": True,
    },
    {
        "name": "Control Type",
        "id": "Control Type",
        "hideable": True,
        "type": "text",
        "editable": True,
    },
    {
        "name": "Sampling Notes",
        "id": "Sampling Notes",
        "hideable": True,
        "type": "text",
        "editable": True,
    },
]

SAMPLE_TABLE_METADATA_COLUMNS = [
    "Assay Name",
    "Assay Date",
    "PP Cert #",
    "Control Type",
    "Sampling Notes",
]

SAMPLE_TABLE_HIDDEN_COLUMNS = [
    "PP Accountable",
    "Sampled Volume",
    "Sampling Device",
    "Sampling Device Type",
    "Processing Technique",
] + SAMPLE_TABLE_METADATA_COLUMNS

PPEL_TABLE_COLUMNS = [
    {"name": "Hardware ID", "id": "Hardware ID", "hideable": True},
    {"name": "Level", "id": "Level", "hideable": True},
    {"name": "Parent ID", "id": "Parent ID", "hideable": True},
    {"name": "Hardware Type", "id": "Hardware Type", "hideable": True},
    {"name": "Dimensionality", "id": "Dimensionality", "hideable": True},
    {"name": "Total Area", "id": "Total Area", "hideable": True},
    {"name": "Total Volume", "id": "Total Volume", "hideable": True},
    {"name": "Analogy", "id": "Analogy", "hideable": True},
    {"name": "Sampled Area", "id": "Sampled Area", "hideable": True},
    {"name": "Sampled Volume", "id": "Sampled Volume", "hideable": True},
    {"name": "Origin", "id": "Origin", "hideable": True},
    {"name": "Spec Class", "id": "Spec Class", "hideable": True},
    {"name": "Spec Value", "id": "Spec Value", "hideable": True},
    {"name": "Grouping", "id": "Grouping", "hideable": True},
    {
        "name": "Grouping Target Density (2D)",
        "id": "Grouping Target Density (2D)",
        "hideable": True,
    },
    {
        "name": "Grouping Target Density (3D)",
        "id": "Grouping Target Density (3D)",
        "hideable": True,
    },
    {"name": "CBE Bioburden Density (2D)", "id": "CBE Bioburden Density (2D)", "hideable": True},
    {"name": "CBE Spore Bioburden (2D)", "id": "CBE Spore Bioburden (2D)", "hideable": True},
    {"name": "CBE Bioburden Density (3D)", "id": "CBE Bioburden Density (3D)", "hideable": True},
    {"name": "CBE Spore Bioburden (3D)", "id": "CBE Spore Bioburden (3D)", "hideable": True},
    {"name": "Handling Constraints", "id": "Handling Constraints", "hideable": True},
    {"name": "Ventilation", "id": "Ventilation", "hideable": True},
    {"name": "Material Composition", "id": "Material Composition", "hideable": True},
    {
        "name": "Cleaning Procedures (Fabrication)",
        "id": "Cleaning Procedures (Fabrication)",
        "hideable": True,
    },
    {
        "name": "Cleaning Procedures (Pre-SI&T)",
        "id": "Cleaning Procedures (Pre-SI&T)",
        "hideable": True,
    },
    {"name": "Cleaning Procedures (SI&T)", "id": "Cleaning Procedures (SI&T)", "hideable": True},
    {
        "name": "Bioburden Reduction (Fabrication)",
        "id": "Bioburden Reduction (Fabrication)",
        "hideable": True,
    },
    {
        "name": "Bioburden Reduction (Pre-SI&T)",
        "id": "Bioburden Reduction (Pre-SI&T)",
        "hideable": True,
    },
    {"name": "Bioburden Reduction (SI&T)", "id": "Bioburden Reduction (SI&T)", "hideable": True},
    {"name": "Hardware Notes", "id": "Hardware Notes", "hideable": True},
]

HARDWARE_METADATA_COLUMNS = [
    "handling",
    "ventilation",
    "composition",
    "cleaning_fab",
    "cleaning_pre",
    "cleaning_sit",
    "reduction_fab",
    "reduction_pre",
    "reduction_sit",
    "notes",
]

PPEL_TABLE_METADATA_COLUMNS = [
    "Handling Constraints",
    "Ventilation",
    "Material Composition",
    "Cleaning Procedures (Fabrication)",
    "Cleaning Procedures (Pre-SI&T)",
    "Cleaning Procedures (SI&T)",
    "Bioburden Reduction (Fabrication)",
    "Bioburden Reduction (Pre-SI&T)",
    "Bioburden Reduction (SI&T)",
    "Hardware Notes",
]

PPEL_TABLE_HIDDEN_COLUMNS = [
    "Parent ID",
    "Dimensionality",
    "Total Volume",
    "Analogy",
    "Sampled Volume",
    "Origin",
    "Spec Class",
    "Spec Value",
    "Grouping",
    "Grouping Target Density (3D)",
    "CBE Bioburden Density (3D)",
    "CBE Spore Bioburden (3D)",
] + PPEL_TABLE_METADATA_COLUMNS

SPEC_DENSITY_MAP = {
    "2D (Area)": {
        "Surface, ISO<=7, Bio Control": 50,
        "Surface, ISO<=7, Particle Control": 500,
        "Surface, ISO=8, Bio Control": 1000,
        "Surface, ISO=8, Particle Control": 10000,
        "Surface, Uncontrolled": 100000,
        "Enclosed Surface, Cleanroom, Particle & Bio Control": 5000,
        "Enclosed Surface, Cleanroom, Only Particle Control": 100000,
        "Enclosed Surface, Uncontrolled Manufacturing": 1000000,
    },
    "3D (Volume)": {
        "Encapsulated, Electronics Piece Parts": 150,
        "Encapsulated, Non-metal Avg": 130,
        "Encapsulated, Non-metal Other": 30,
    },
}

TABLE_DEF_PPEL = [
    [
        "Hardware ID",
        "Unique identifier for this hardware element. Can be any string but descriptive/semantic names are typically better since this ID will be used to reference the element throughout the tool.",
        "N/A",
        "String",
        "Unique",
        "RI",
    ],
    [
        "Level",
        "Rollup level of the hardware element, indicating level of nesting in the hardware hierarchy (= level of parent + 1). Level 1 denotes the project-wide top level.",
        "N/A",
        "Integer",
        "x >= 1",
        "D",
    ],
    [
        "Parent ID",
        "ID of this element's parent hardware in the hierarchy (if any). This field alongside Hardware ID uniquely defines the overall hardware structure of the project. Elements without a parent will be assumed to be Level 2 (top level beneath project root).",
        "N/A",
        "String",
        "Valid Hardware ID from Hierarchy",
        "RI",
    ],
    [
        "Hardware Type",
        "For rollups, a simple flag to indicate this this hardware element has children. For components (leaf nodes in the hardware tree), this is the basis for the bioburden estimate. If the component was sampled, then the sampling event observation is used to estimate the bioburden. If it was not sampled, then the bioburden can be estimated using the estimate from another component (it is “implied” from that component) of the spec value.",
        "N/A",
        "String",
        "['Rollup', 'Sampled', 'Unsampled - Implied', 'Unsampled - Spec']",
        "RI",
    ],
    [
        "Dimensionality",
        "Supplied for component elements only (leaf nodes in the hardware tree). Indicates a 2D component (estimating surface area bioburden) or a 3D component (estimating enclosed volume bioburden).",
        "N/A",
        "String",
        "['2D (Area)', '3D (Volume)']",
        "RI",
    ],
    [
        "Total Area",
        "Total surface area of the component being sampled. Required input for components (leaf nodes), derived for rollups based on aggregation over component children. This field is left blank if the component is not an area but a volume.",
        "m²",
        "Float",
        "x > 0",
        "RI / D",
    ],
    [
        "Total Volume",
        "Total volume of the component being sampled. Required input for components (leaf nodes), derived for rollups based on aggregation over component children. This field is left blank if the component is not a volume but an area.",
        "cm³",
        "Float",
        "x > 0",
        "RI / D",
    ],
    [
        "Analogy",
        "ID of a known or estimated analog, determining the bioburden density prior for this component before observing any samples. '-- Generic --' for noninformative prior. Only relevant for sampled components.",
        "N/A",
        "String",
        "Valid Analog from Tool Database",
        "RI",
    ],
    [
        "Sampled Area",
        "Sum of areas for all accountable samples for this hardware element (for components, sum over samples linked to this element, or blank if component is 3D; for rollups, sum over samples linked to all 2D component children).",
        "m²",
        "Float",
        "x >= 0",
        "D",
    ],
    [
        "Sampled Volume",
        "Sum of volumes for all accountable samples for this hardware element (for components, sum over samples linked to this element, or blank if component is 2D; for rollups, sum over samples linked to all 3D component children).",
        "cm³",
        "Float",
        "x >= 0",
        "D",
    ],
    [
        "Origin",
        "ID of another component in the hierarchy, whose bioburden will also be assumed for this hardware as well. Only relevant for implied components.",
        "N/A",
        "String",
        "Valid Hardware ID from Hierarchy",
        "RI",
    ],
    [
        "Spec Class",
        "Facility class for the hardware, which determines the assumed bioburden for hardware that is not samples or implied. Only relevant for spec components.",
        "N/A",
        "String",
        "For 2D Components: ['Surface, ISO<=7, Bio Control', 'Surface, ISO<=7, Particle Control', 'Surface, ISO=8, Bio Control', 'Surface, ISO=8, Particle Control', 'Surface, Uncontrolled', 'Enclosed Surface, Cleanroom, Particle & Bio Control', 'Enclosed Surface, Cleanroom, Only Particle Control', 'Enclosed Surface, Uncontrolled Manufacturing'], For 3D Components: ['Encapsulated, Electronics Piece Parts', 'Encapsulated, Non-metal Avg', 'Encapsulated, Non-metal Other']",
        "RI",
    ],
    [
        "Spec Value",
        "Uniquely determined by spec class based on NASA STD 8719.27. Only relevant for spec components.",
        "N/A",
        "String",
        "None",
        "RI",
    ],
    [
        "Grouping",
        "A name given to a set of hardware components that share a common threshold bioburden density requirement.",
        "N/A",
        "String",
        "None",
        "RI",
    ],
    [
        "Grouping Target Density (2D)",
        "The surface area spore density threshold not to be exceeded by the 2D hardware components in a given grouping.",
        "spores / m²",
        "Float",
        "x > 0",
        "RI",
    ],
    [
        "Grouping Target Density (3D)",
        "The volume spore density threshold not to be exceeded by the 3D hardware components in a given grouping.",
        "spores / cm³",
        "Float",
        "x > 0",
        "RI",
    ],
    [
        "CBE Bioburden Density (2D)",
        "Single value summary of the estimate for the distribution of surface bioburden density, as calculated by the Bayesian model.",
        "spores / m²",
        "Float",
        "x > 0",
        "D",
    ],
    [
        "CBE Bioburden Density (3D)",
        "Single value summary of the estimate for the distribution of volume bioburden density, as calculated by the Bayesian model.",
        "spores / cm³",
        "Float",
        "x > 0",
        "D",
    ],
    [
        "CBE Spore Bioburden (2D)",
        "Single value summary of the estimate for the distribution of surface bioburden density, as calculated by the Bayesian model.",
        "spores",
        "Float",
        "x > 0",
        "D",
    ],
    [
        "CBE Spore Bioburden (3D)",
        "Single value summary of the estimate for the distribution of volume bioburden density, as calculated by the Bayesian model.",
        "spores",
        "Float",
        "x > 0",
        "D",
    ],
    [
        "Handling Constraints",
        "Text briefly describing what cannot be cleaned or directly sampled.",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
    [
        "Ventilation",
        "Brief description of how the component is vented (e.g., hermetically sealed).",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
    [
        "Material Composition",
        "Brief description of the material type (e.g., metallic, PWB components, Al 6061).",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
    [
        "Cleaning Procedures (Fabrication)",
        "Brief description of the cleaning procedures used during fabrication of the component (e.g., IPA wipe, alcohol rinse, precision cleaning, surface DHMR).",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
    [
        "Cleaning Procedures (Pre-SI&T)",
        "Brief description of the cleaning procedures used after fabrication but before Systems Integration and Test (SI&T) of the component (e.g., IPA wipe, bulk DHMR, surface DHMR).",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
    [
        "Cleaning Procedures (SI&T)",
        "Brief description of the cleaning procedures used during Systems Integration and Test (SI&T) of the component (e.g., IPA wipe, bulk DHMR, surface DHMR).",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
    [
        "Bioburden Reduction (Fabrication)",
        "Orders of magnitude reduction to bioburden from fabrication cleaning procedures.",
        "unitless",
        "Float",
        "x >= 0",
        "MDO",
    ],
    [
        "Bioburden Reduction (Pre-SI&T)",
        "Orders of magnitude reduction to bioburden from pre-SI&T cleaning procedures.",
        "unitless",
        "Float",
        "x > 0",
        "MDO",
    ],
    [
        "Bioburden Reduction (SI&T)",
        "Orders of magnitude reduction to bioburden from SI&T cleaning procedures.",
        "unitless",
        "Float",
        "x > 0",
        "MDO",
    ],
    [
        "Hardware Notes",
        "Any other details that should be documented for this component not captured in other fields.",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
]

TABLE_DEF_SAMPLES = [
    [
        "Sample ID",
        "Unique identifier for this sample. Can be any string but descriptive/semantic names are typically better since this ID will be used to reference the sample throughout the tool.",
        "N/A",
        "String",
        "Unique",
        "RI",
    ],
    [
        "Hardware ID",
        "ID of the hardware element that was sampled. All samples must have one and only one Hardware ID, but each hardware element may be associated with many samples. Only hardware elements with Component Type = 'Sampled' are eligible for this field.",
        "N/A",
        "String",
        "Valid Hardware ID from Hierarchy",
        "RI",
    ],
    [
        "PP Accountable",
        "Indicates whether or not the sample should be considered in the bioburden estimate calculations (“Yes” = include in the bioburden estimate; “No” = Do not include in the bioburden estimate).",
        "N/A",
        "String",
        "['Yes', 'No']",
        "RI",
    ],
    [
        "Sampled Area",
        "Area of the surface to which the sampling device was applied. Only relevant for samples on 2D hardware elements.",
        "m²",
        "Float",
        "x > 0",
        "RI",
    ],
    [
        "Sampled Volume",
        "Volume of the region to which the sampling device was applied. Only relevant for samples on 3D hardware elements.",
        "cm³",
        "Float",
        "x > 0",
        "RI",
    ],
    [
        "Sampling Device",
        "The family of devices used for sampling (i.e., swab, wipe). Determines the options available for Sampling Device Type.",
        "N/A",
        "String",
        "['Swab', 'Wipe']",
        "RI",
    ],
    [
        "Sampling Device Type",
        "The specific kind of device used for sampling (the specific type of swab or wipe). Used in determining the recovery efficiency distribution to assume for this sample.",
        "N/A",
        "String",
        "For Swabs: ['Puritan Cotton', 'Nylon-flocked', 'Copan Polyester', 'Copan Cotton'], For Wipes: ['TX3211', 'TX3224']",
        "RI",
    ],
    [
        "Processing Technique",
        "The assay process applied during the sampling event. Used in determining the recovery efficiency distribution to assume for this sample.",
        "N/A",
        "String",
        "['NASA Standard', 'NASA Standard (w/ Membrane Filtration)', 'ESA Standard', 'ESA Standard (w/ Membrane Filtration)']",
        "RI",
    ],
    [
        "Pour Fraction",
        "The fraction of the solution containing the sample that is plated on a petri dish for growth and CFU observation. Default values are 0.8 for swab-pour plate samples; 0.25 for wipe-pour plate samples; and 0.92 for sampling events using membrane filtration.",
        "unitless",
        "Float",
        "0 < x <= 1",
        "RI",
    ],
    [
        "CFU",
        "The sum of all colony forming units (CFU) observed across all petri dishes used for the sample after 72 hours incubation time.",
        "spores",
        "Integer",
        "x >= 0",
        "RI",
    ],
    [
        "Assay Name",
        "Name or code given to the assay in which the sample was taken.",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
    [
        "Assay Date",
        "Year, month and day that the sample was taken.",
        "N/A",
        "Date (ISO-8601)",
        "None",
        "MDO",
    ],
    [
        "PP Cert #",
        "Number certifying the final sampling of a component.",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
    [
        "Control Type",
        "If the sample is a control sample, this captures the kind of control. If the type of control is not listed, describe it in the Sampling Notes field. The default is 'Not Control'.",
        "N/A",
        "String",
        "['Facility Control', 'Negative Control', 'Positive Control', 'Field Control', 'Other Control', 'Not Control']",
        "MDO",
    ],
    [
        "Sampling Notes",
        "Any other details that should be documented for this sample not captured in other fields.",
        "N/A",
        "String",
        "None",
        "MDO",
    ],
]


# Configure recovery efficiency distributions for each device/type/technique combination
# Params obtained from simple minimization algorithm in scripts/fit_efficiency_beta.py
# For more details on how these distributions were generated see docs/fit_efficiency_beta.md
EFFICIENCY_CONFIG = {
    "Swab;Puritan Cotton;NASA Standard": {
        "params": [45.56431672969219, 100.24149680532281],
        "default_fraction": 0.8,
    },
    "Swab;Puritan Cotton;NASA Standard (w/ Membrane Filtration)": {
        "params": [97.55218540553831, 191.9575261205754],
        "default_fraction": 0.92,
    },
    "Swab;Puritan Cotton;ESA Standard": {
        "params": "Swab;Puritan Cotton;NASA Standard",
        "default_fraction": 0.8,
    },
    "Swab;Puritan Cotton;ESA Standard (w/ Membrane Filtration)": {
        "params": "Swab;Puritan Cotton;NASA Standard (w/ Membrane Filtration)",
        "default_fraction": 0.92,
    },
    "Swab;Nylon-flocked;NASA Standard": {
        "params": [9.579630660559655, 23.74082381095219],
        "default_fraction": 0.8,
    },
    "Swab;Nylon-flocked;NASA Standard (w/ Membrane Filtration)": {
        "params": "Swab;Nylon-flocked;NASA Standard",
        "default_fraction": 0.92,
    },
    "Swab;Nylon-flocked;ESA Standard": {
        "params": [68.16498856079723, 75.34025051456537],
        "default_fraction": 0.8,
    },
    "Swab;Nylon-flocked;ESA Standard (w/ Membrane Filtration)": {
        "params": "Swab;Nylon-flocked;ESA Standard",
        "default_fraction": 0.92,
    },
    "Swab;Copan Polyester;NASA Standard": {
        "params": "Swab;Copan Polyester;ESA Standard",
        "default_fraction": 0.8,
    },
    "Swab;Copan Polyester;NASA Standard (w/ Membrane Filtration)": {
        "params": "Swab;Copan Polyester;ESA Standard",
        "default_fraction": 0.92,
    },
    "Swab;Copan Polyester;ESA Standard": {
        "params": [6.052080310455172, 42.3645621731862],
        "default_fraction": 0.8,
    },
    "Swab;Copan Polyester;ESA Standard (w/ Membrane Filtration)": {
        "params": "Swab;Copan Polyester;ESA Standard",
        "default_fraction": 0.92,
    },
    "Swab;Copan Cotton;NASA Standard": {
        "params": [51.836071542660086, 362.8525007986206],
        "default_fraction": 0.8,
    },
    "Swab;Copan Cotton;NASA Standard (w/ Membrane Filtration)": {
        "params": "Swab;Copan Cotton;NASA Standard",
        "default_fraction": 0.92,
    },
    "Swab;Copan Cotton;ESA Standard": {
        "params": "Swab;Copan Cotton;NASA Standard",
        "default_fraction": 0.8,
    },
    "Swab;Copan Cotton;ESA Standard (w/ Membrane Filtration)": {
        "params": "Swab;Copan Cotton;NASA Standard",
        "default_fraction": 0.92,
    },
    "Wipe;TX3211;NASA Standard": {
        "params": "Wipe;TX3211;NASA Standard (w/ Membrane Filtration)",
        "default_fraction": 0.25,
    },
    "Wipe;TX3211;NASA Standard (w/ Membrane Filtration)": {
        "params": [2.755428498737132, 7.13349822450835],
        "default_fraction": 0.92,
    },
    "Wipe;TX3211;ESA Standard": {
        "params": "Wipe;TX3211;NASA Standard (w/ Membrane Filtration)",
        "default_fraction": 0.25,
    },
    "Wipe;TX3211;ESA Standard (w/ Membrane Filtration)": {
        "params": "Wipe;TX3211;NASA Standard (w/ Membrane Filtration)",
        "default_fraction": 0.92,
    },
    "Wipe;TX3224;NASA Standard": {
        "params": "Wipe;TX3224;NASA Standard (w/ Membrane Filtration)",
        "default_fraction": 0.25,
    },
    "Wipe;TX3224;NASA Standard (w/ Membrane Filtration)": {
        "params": [38.27721767664384, 259.32814975926203],
        "default_fraction": 0.92,
    },
    "Wipe;TX3224;ESA Standard": {
        "params": "Wipe;TX3224;NASA Standard (w/ Membrane Filtration)",
        "default_fraction": 0.25,
    },
    "Wipe;TX3224;ESA Standard (w/ Membrane Filtration)": {
        "params": "Wipe;TX3224;NASA Standard (w/ Membrane Filtration)",
        "default_fraction": 0.92,
    },
}


def get_efficiency_params(tag: str) -> list[float, float]:
    """Fetch the appropriate beta parameters for efficiency distribution"""

    params = EFFICIENCY_CONFIG[tag]["params"]
    return EFFICIENCY_CONFIG[params]["params"] if isinstance(params, str) else params


def resolve_hardware_categorical(val: str):
    """Attempt to resolve user input into a categorical value for hardware"""

    val = "" if val is None else val

    if "electronics piece parts" in val.lower():
        return "Encapsulated, Electronics Piece Parts"
    elif "non-metal avg" in val.lower():
        return "Encapsulated, Non-metal Avg"
    elif "non-metal other" in val.lower():
        return "Encapsulated, Non-metal Other"
    elif "iso<=7" in val.lower() and "bio control" in val.lower():
        return "Surface, ISO<=7, Bio Control"
    elif "iso<=7" in val.lower() and "particle control" in val.lower():
        return "Surface, ISO<=7, Particle Control"
    elif "iso=8" in val.lower() and "bio control" in val.lower():
        return "Surface, ISO=8, Bio Control"
    elif "iso=8" in val.lower() and "particle control" in val.lower():
        return "Surface, ISO=8, Particle Control"
    elif "enclosed" not in val.lower() and "uncontrolled" in val.lower():
        return "Surface, Uncontrolled"
    elif "enclosed" in val.lower() and "particle" in val.lower() and "bio" in val.lower():
        return "Enclosed Surface, Cleanroom, Particle & Bio Control"
    elif "enclosed" in val.lower() and "particle" in val.lower():
        return "Enclosed Surface, Cleanroom, Only Particle Control"
    elif "enclosed" in val.lower() and "uncontrolled" in val.lower():
        return "Enclosed Surface, Uncontrolled Manufacturing"

    if val.startswith("2"):
        return "2D (Area)"
    elif val.startswith("3"):
        return "3D (Volume)"

    val_std = val.lower()
    if "implied" in val_std:
        return "Unsampled - Implied"
    elif "spec" in val_std:
        return "Unsampled - Spec"
    elif "sampled" in val_std:
        return "Sampled"
    elif "rollup" in val_std:
        return "Rollup"

    return ""


def resolve_hardware_analogy(val: str):
    """Attempt to resolve user input into an analogy for a sampled component"""

    return val if val in POSTERIOR_MAP else ""


def resolve_sample_categorical(val: str):
    """Attempt to resolve user input into a categorical value for samples"""

    categorical_map = {
        "swab": "Swab",
        "wipe": "Wipe",
        "puritan cotton": "Puritan Cotton",
        "nylon flocked": "Nylon-flocked",
        "copan cotton": "Copan Cotton",
        "copan polyester": "Copan Polyester",
        "tx3211": "TX3211",
        "tx3224": "TX3224",
        "nasa standard (w/ membrane filtration)": "NASA Standard (w/ Membrane Filtration)",
        "nasa standard": "NASA Standard",
        "esa standard (w/ membrane filtration)": "ESA Standard (w/ Membrane Filtration)",
        "esa standard": "ESA Standard",
    }
    val = "" if val is None else val
    val_std = val.replace("-", " ").lower().strip()

    return categorical_map[val_std] if val_std in categorical_map else ""


def load_semantic_map_json(semantic_map_path: str) -> dict | None:
    """Load pickle file containing maps of analogy IDs from Andrei's to a semantic name

    Pass invalid file name (e.g. empty string) to disable semantic map.
    """

    if not os.path.exists(semantic_map_path) or os.path.isdir(semantic_map_path):
        return None

    with open(semantic_map_path, "r") as f:
        return json.load(f)


def load_posterior_samples_json(
    posterior_path: str, semantic_map_path: str
) -> dict[str, np.ndarray]:
    """Load a mapping of components to posterior samples"""

    semantic_map = load_semantic_map_json(semantic_map_path)

    with open(posterior_path, "r") as f:
        posterior_dict = json.load(f)

    if semantic_map:
        posterior_dict = {semantic_map.get(k, k): np.array(v) for k, v in posterior_dict.items()}
    else:
        posterior_dict = {k: np.array(v) for k, v in posterior_dict.items()}

    return posterior_dict


def create_component_tree(
    hierarchy_path: str,
    semantic_map_path: str,
    rollup_cols: list[str],
    mode_volumes: bool = False,
) -> dict:
    """Create a hierarchical structure representing the nested components from project hierarchy"""

    # Load base hierarchy
    df_hier = pd.read_excel(hierarchy_path, engine="openpyxl")
    df_hier = df_hier[df_hier["Component Type"] == "Sampled"]
    df_hier = df_hier[df_hier["Is Volume"]] if mode_volumes else df_hier[~df_hier["Is Volume"]]

    # Load semantic map
    semantic_map = load_semantic_map_json(semantic_map_path)

    # Create a recursive defaultdict for automatic nested dictionary creation
    def nested_dict():
        return defaultdict(nested_dict)

    # Build the tree
    tree = nested_dict()
    for i in range(len(df_hier.index)):
        current = tree
        for col in rollup_cols:
            value = df_hier.loc[i, col]
            if not isinstance(value, str):
                break
            current = current[str(value)]

    # Format tree for dash-mantine-components
    def format_tree(subtree, path=""):
        result = []

        for value, children in sorted(subtree.items()):
            # Create the full path for the value
            current_path = f"{path}/{value}" if path else value

            # Create a node
            node = {
                "value": current_path,
                "label": semantic_map.get(value, value) if semantic_map else value,
            }

            # If the node has children, add them recursively
            if children:
                node["children"] = format_tree(children, current_path)

            result.append(node)

        return result

    return format_tree(tree), list(df_hier[rollup_cols[0]].unique())


def component_tree_entries(tree, skip=[]):
    """Identify all unique components in the component tree"""

    labels = set()
    for node in tree:
        labels.add(node.get("label"))
        labels.update(component_tree_entries(node.get("children", [])))
    return labels - set(skip)


def find_by_key(arr: list[dict], key: str, value: str):
    """Find first element in list of dicts with matching value at given key"""

    arr_filt = [x for x in arr if x[key] == value]
    return None if not arr_filt else arr_filt[0]


def parse_ppel_upload(contents: str, filename: str):
    """Convert uploaded PPEL file into a list of hardware elements or return error message"""

    # Prefine the columns we want to extract
    cols_ppel = [
        "Hardware ID",
        "Level",
        "Parent ID",
        "Hardware Type",
        "Dimensionality",
        "Total Area",
        "Total Volume",
        "Analogy",
        "Origin",
        "Spec Class",
        "Grouping",
        "Grouping Target Density (2D)",
        "Grouping Target Density (3D)",
    ] + PPEL_TABLE_METADATA_COLUMNS
    cols_samples = [
        "Sample ID",
        "Hardware ID",
        "PP Accountable",
        "Sampled Area",
        "Sampled Volume",
        "Sampling Device",
        "Sampling Device Type",
        "Processing Technique",
        "Pour Fraction",
        "CFU",
    ] + SAMPLE_TABLE_METADATA_COLUMNS

    # Decode the uploaded contents
    _, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)
    pseudofile = io.BytesIO(decoded)

    # Ensure correct file type uploaded
    if not filename.endswith(".xlsx"):
        return False, "Upload must be an Excel file with .xlsx extension"

    # Attempt to parse upload into dataframes
    try:
        df_ppel = pd.read_excel(pseudofile, sheet_name="PPEL")
        df_ppel = df_ppel.map(lambda x: None if pd.isna(x) else x)
        df_samples = pd.read_excel(pseudofile, sheet_name="Samples")
        df_samples = df_samples.map(lambda x: None if pd.isna(x) else x)
    except Exception as _:
        return False, "Error unpacking file contents, ensure Excel file is not corrupted"

    # Extract just the columns we need - PPEL sheet
    try:
        df_ppel = df_ppel[cols_ppel].copy()
    except Exception as _:
        return (
            False,
            f"Error extracting target columns from PPEL worksheet, ensure all of the following columns exist: {cols_ppel}",
        )

    # Extract just the columns we need - Samples sheet
    try:
        df_samples = df_samples[cols_samples].copy()
    except Exception as _:
        return (
            False,
            f"Error extracting target columns from Samples worksheet, ensure all of the following columns exist: {cols_samples}",
        )

    # Filter out the project row and extract project name & group tag
    try:
        project_row = df_ppel[df_ppel["Level"] == 1].loc[0]
        data_project = {"name": project_row["Hardware ID"], "group": project_row["Grouping"]}
        df_ppel = df_ppel[df_ppel["Level"] != 1]
    except Exception as _:
        return (
            False,
            "Error identifying project-level row from cells, ensure correct format is being used",
        )

    # Ensure no duplicate hardware ids
    if len(df_ppel["Hardware ID"]) != len(df_ppel["Hardware ID"].unique()):
        return False, "Upload has duplicate 'Hardware ID' entries, resolve this before uploading"

    # Convert to json and resolve any categoricals
    data_hardware = df_ppel.to_dict("records")
    unresolved_hardware = []
    for record, f in product(
        data_hardware,
        [
            "Dimensionality",
            "Hardware Type",
            "Spec Class",
            "Analogy",
        ],
    ):
        resolved = (
            resolve_hardware_categorical(record[f])
            if f != "Analogy"
            else resolve_hardware_analogy(record[f])
        )
        if record[f] and not resolved:
            unresolved_hardware.append((record["Hardware ID"], f))
        record[f] = resolved
    data_samples = df_samples.to_dict("records")
    unresolved_samples = []
    for record, f in product(
        data_samples, ["Sampling Device", "Sampling Device Type", "Processing Technique"]
    ):
        resolved = resolve_sample_categorical(record[f])
        if record[f] and not resolved:
            unresolved_samples.append((record["Sample ID"], f))
        record[f] = resolved

    return True, (
        data_hardware,
        unresolved_hardware,
        data_samples,
        unresolved_samples,
        data_project,
    )


def parse_pps_upload(contents: str, filename: str):
    """Convert uploaded PPS file into list of samples or return error message"""

    def extract_raw_df(df_raw, pp_acc):
        assay_name = df_raw.loc[2, 2]
        assay_date = datetime.strptime(df_raw.loc[2, 3], "%m/%d/%Y %H:%M:%S").date().isoformat()
        idx_samples_header = list(df_raw.loc[:, 0]).index("Sample Number")
        df = pd.DataFrame(df_raw.loc[idx_samples_header + 1 :, :6].to_dict())
        df.columns = df_raw.loc[9, :6]
        df["Assay Name"] = assay_name
        df["Assay Date"] = assay_date
        df["PP Accountable"] = pp_acc
        return df

    # Decode the uploaded contents
    _, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)
    pseudofile = io.BytesIO(decoded)

    # Ensure correct file type uploaded
    if not filename.endswith(".xlsx"):
        return False, "Upload must be an Excel file with .xlsx extension"

    # Attempt to parse upload into dataframe(s)
    try:
        wb = load_workbook(pseudofile)
        if not all([n in ["Accountable", "Non-Accountable"] for n in wb.sheetnames]):
            return (
                False,
                "Excel file should only contain sheet(s) named 'Accountable' or 'Non-Accountable'",
            )
        acc_key = {"Accountable": "Yes", "Non-Accountable": "No"}
        dfs_raw = [
            (pd.read_excel(pseudofile, header=None, sheet_name=n), acc_key[n])
            for n in wb.sheetnames
        ]
    except Exception as _:
        return False, "Error unpacking file contents, ensure Excel file is not corrupted"

    # Extract the sample rows
    try:
        df = pd.concat([extract_raw_df(raw, acc) for (raw, acc) in dfs_raw])
    except Exception as _:
        return False, "Error extracting samples from cells, ensure correct format is being used"

    # Ensure no duplicate sample numbers
    if len(df["Sample Number"]) != len(df["Sample Number"].unique()):
        return False, "Upload has duplicate 'Sample Number' entries, resolve this before uploading"

    # Convert to json and resolve any categoricals
    data_samples = df.to_dict("records")
    unresolved_samples = []
    pps_field_map = {"Sampling Method": "Sampling Device"}
    for record, f in product(data_samples, ["Sampling Method"]):
        resolved = resolve_sample_categorical(record[f])
        if record[f] and not resolved:
            unresolved_samples.append((record["Sample Number"], pps_field_map[f]))
        record[f] = resolved

    # Add default metadata
    for record in data_samples:
        record["PP Cert #"] = ""
        record["Control Type"] = "Not Control"
        record["Sampling Notes"] = ""

    return True, (data_samples, unresolved_samples)


POSTERIOR_MAP = load_posterior_samples_json(PATH_POSTERIOR, PATH_SEMANTIC_MAP)

ANALOGY_TREE_AREAS, L1_PROJECTS_AREAS = create_component_tree(
    PATH_HIERARCHY,
    PATH_SEMANTIC_MAP,
    [
        "Rollup Level 1: Project",
        "Rollup Level 2: Payload / Spacecraft",
        "Rollup Level 3: Instrument / Flight Element",
        "Rollup Level 4: Assembly",
        "Rollup Level 5: Subassembly",
        "Rollup Level 6: Component",
    ],
    False,
)

ANALOGY_TREE_VOLUMES, L1_PROJECTS_VOLUMES = create_component_tree(
    PATH_HIERARCHY,
    PATH_SEMANTIC_MAP,
    [
        "Rollup Level 1: Project",
        "Rollup Level 2: Payload / Spacecraft",
        "Rollup Level 3: Instrument / Flight Element",
        "Rollup Level 4: Assembly",
        "Rollup Level 5: Subassembly",
        "Rollup Level 6: Component",
    ],
    True,
)

ANALOGY_COMPONENTS_AREAS = sorted(
    list(component_tree_entries(ANALOGY_TREE_AREAS, skip=L1_PROJECTS_AREAS)),
    # Uncomment below if using "Component X" names to force sort ascending
    # key=lambda x: int(x[x.index(" ") + 1 :]),
)
ANALOGY_COMPONENTS_VOLUMES = sorted(
    list(component_tree_entries(ANALOGY_TREE_VOLUMES, skip=L1_PROJECTS_VOLUMES)),
    # Uncomment below if using "Component X" names to force sort ascending
    # key=lambda x: int(x[x.index(" ") + 1 :]),
)

with open(PATH_METADATA, "r") as f:
    ANALOGY_METADATA = json.load(f)
